import os
import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter
import xlsxwriter

def Merge_Files(FILE_DIR,OUTPUT_FILENAME):
    # Read input files
    cwd = os.path.join(FILE_DIR)
    files = os.listdir(cwd)
    # Create Blank Master Excel Workbook
    workbook = xlsxwriter.Workbook(OUTPUT_FILENAME)
    worksheet = workbook.add_worksheet()
    workbook.close()
    # Load excel Workbook using openpyxl
    book = load_workbook(OUTPUT_FILENAME)
    writer = ExcelWriter(OUTPUT_FILENAME, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # create empty DataFrame
    df = pd.DataFrame()
    for file in files:
        if file.endswith('.xlsx'):
            sheetname = file.replace(".xlsx", "")
            df = pd.read_excel(os.path.join(FILE_DIR, file), sheet_name=sheetname)
            df.to_excel(writer, sheet_name=sheetname, index=False, header=True)
        if file.endswith('.csv'):
            sheetname = file.replace(".csv", "")
            df = pd.read_csv(os.path.join(FILE_DIR, file))
            df.to_excel(writer, sheet_name=sheetname, index=False, header=True)
    # Remove empty sheet from Master.xlsx
    first_sheet = book['Sheet1']
    book.remove(first_sheet)
    writer.save()
    print("MASTER FILE GENERATED")

# call function to read multiple csv's to generate MASTER EXCEL FILE
# Merge_Files("CSV_FILES","MASTER_CSV.xlsx")
# call function to read multiple excel's to generate MASTER EXCEL FILE
Merge_Files("EXCEL_FILES","MASTER_EXCEL.xlsx")